"""
CANF_OTPP main — orchestrates the full pipeline.

    scraper.fetch_data()  →  mapper.map_to_output()  →  DATA / META / ZIP

Usage:
    python main.py                                          # download + process
    python main.py --skip-download --source-pdf <path>     # skip download
"""

import argparse
import os
import re
import shutil
import zipfile
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Font

import scraper
import mapper

OUTPUT_PREFIX = "CANF_OTPP"
DOWNLOADS_DIR = "downloads"
OUTPUT_DIR    = "output"


def _datestamp() -> str:
    return datetime.now().strftime('%Y%m%d')


def _apply_number_format(filepath: str):
    """Apply comma format to all numeric data cells (rows 3+, cols 2+)."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    for row in ws.iter_rows(min_row=3, min_col=2):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.##"
    wb.save(filepath)


def _save_data(df: pd.DataFrame, datestamp: str) -> str:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    path = os.path.join(OUTPUT_DIR, f"{OUTPUT_PREFIX}_DATA_{datestamp}.xlsx")
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='DATA', index=False, header=False)
        ws = writer.sheets['DATA']
        bold = Font(bold=True)
        for col in ws.iter_cols(min_row=1, max_row=2):
            for cell in col:
                cell.font = bold
        for col in ws.columns:
            width = max(len(str(c.value or '')) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(width + 2, 50)
    _apply_number_format(path)
    print(f"[OK] DATA  -> {path}")
    return path


def _save_metadata(datestamp: str) -> str:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    path = os.path.join(OUTPUT_DIR, f"{OUTPUT_PREFIX}_META_{datestamp}.xlsx")
    pd.DataFrame(mapper.build_metadata_rows()).to_excel(path, index=False)
    print(f"[OK] META  -> {path}")
    return path


def _create_zip(data_path: str, meta_path: str, datestamp: str) -> str:
    zip_path = os.path.join(OUTPUT_DIR, f"{OUTPUT_PREFIX}_{datestamp}.ZIP")
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.write(data_path, os.path.basename(data_path))
        zf.write(meta_path, os.path.basename(meta_path))
    print(f"[OK] ZIP   -> {zip_path}")
    return zip_path


def scrape(skip_download: bool = False, source_pdf: str = None):
    """Run the full pipeline and write DATA, META, ZIP to output/."""
    os.makedirs(DOWNLOADS_DIR, exist_ok=True)
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    if skip_download and source_pdf:
        src = Path(source_pdf)
        dest = os.path.join(DOWNLOADS_DIR, src.name)
        if not os.path.exists(dest):
            shutil.copy2(source_pdf, dest)

        # Infer report type and year from filename.
        # Strip leading CANF_OTPP_DATA_YYYYMMDD_ prefix so the datestamp
        # doesn't shadow the actual report year embedded in the filename.
        name = re.sub(r'^canf_otpp_data_\d{8}_', '', src.stem.lower())
        m = re.search(r'(20\d{2})', name)
        year = int(m.group(1)) if m else datetime.now().year
        rtype = 'annual' if 'annual' in name else 'interim'
        report_info = {'type': rtype, 'year': year}

        start, end = scraper.find_schedule_pages(dest)
        rows = scraper.parse_table(dest, start, end)
        raw_df = pd.DataFrame(rows, columns=['label', 'fv_current', 'cost_current', 'fv_prior', 'cost_prior'])
        print(f"[OK] Parsed {len(raw_df)} rows from {src.name}")
    else:
        raw_df, report_info = scraper.fetch_data(DOWNLOADS_DIR)

    print(f"[INFO] Report: {report_info['type'].title()} {report_info['year']}")
    output_df = mapper.map_to_output(raw_df, report_info)

    ds = _datestamp()
    data_path = _save_data(output_df, ds)
    meta_path = _save_metadata(ds)
    _create_zip(data_path, meta_path, ds)
    print(f"\n[DONE] {report_info['type'].title()} {report_info['year']} — output in {OUTPUT_DIR}/")


def main():
    parser = argparse.ArgumentParser(description="CANF_OTPP pipeline")
    parser.add_argument('--skip-download', action='store_true',
                        help="Skip browser download step")
    parser.add_argument('--source-pdf', default=None,
                        help="Path to an existing PDF to process")
    args = parser.parse_args()
    scrape(skip_download=args.skip_download, source_pdf=args.source_pdf)


if __name__ == '__main__':
    main()
