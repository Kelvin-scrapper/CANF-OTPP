import pandas as pd
import numpy as np
from datetime import datetime
import openpyxl
import os

class OTTPHierarchicalExtractor:
    """
    OTPP Data Extractor with full logic restored.
    This version includes the FINAL ROW OFFSET FIX, adjusting all mappings
    by +10 rows to match the actual file structure.
    """
    
    def __init__(self):
        # ==============================================================================
        # MAPPING LOGIC CORRECTED with +10 ROW OFFSET to match the actual file.
        # This is the definitive mapping.
        # Liability row numbers (53+) are best guesses and should be verified.
        # ==============================================================================
        self.row_mappings = {
            # === EQUITY SECTION (Fixed to actual data rows) ===
            19: {'asset_code': 'EQUITY', 'context': 'Equity Total (after Non-Canadian private equity)', 'description': 'OTPP Investments, Fair Value, Equity', 'cost_description': 'OTPP Investments, Cost, Equity'},
            12: {'asset_code': 'EQUITYHEADER', 'context': 'Equity (Header Row - produces blank column)', 'description': 'OTPP Investments, Fair Value, Equity Header', 'cost_description': 'OTPP Investments, Cost, Equity Header'},
            13: {'asset_code': 'EQUITIES', 'context': 'Publicly Traded (Header Row - produces blank column)', 'description': 'OTPP Investments, Fair Value, Publicly Traded', 'cost_description': 'OTPP Investments, Cost, Publicly Traded'},
            14: {'asset_code': 'DOMESTICEQUITIES', 'context': 'Canadian under Publicly traded', 'description': 'OTPP Investments, Fair Value, Publicly Traded, Canadian', 'cost_description': 'OTPP Investments, Cost, Publicly Traded, Canadian'},
            15: {'asset_code': 'FOREIGNEQUITIES', 'context': 'Non-Canadian under Publicly traded', 'description': 'OTPP Investments, Fair Value, Publicly Traded, Non-Canadian', 'cost_description': 'OTPP Investments, Cost, Publicly Traded, Non-Canadian'},
            16: {'asset_code': 'PRIVATEEQUITY', 'context': 'Non-publicly traded (Header Row - produces blank column)', 'description': 'OTPP Investments, Fair Value, Non-publicly traded', 'cost_description': 'OTPP Investments, Cost, Non-publicly traded'},
            17: {'asset_code': 'DOMESTICPRIVATEEQUITY', 'context': 'Canadian under Non-publicly traded', 'description': 'OTPP Investments, Fair Value, Non-publicly traded, Canadian', 'cost_description': 'OTPP Investments, Cost, Non-publicly traded, Canadian'},
            18: {'asset_code': 'FOREIGNPRIVATEEQUITY', 'context': 'Non-Canadian under Non-publicly traded', 'description': 'OTPP Investments, Fair Value, Non-publicly traded, Non-Canadian', 'cost_description': 'OTPP Investments, Cost, Non-publicly traded, Non-Canadian'},

            # === FIXED INCOME SECTION (Fixed to actual data rows) ===
            26: {'asset_code': 'FIXEDINCOME', 'context': 'Fixed Income Total (after Other debt)', 'description': 'OTPP Investments, Fair Value, Fixed Income', 'cost_description': 'OTPP Investments, Cost, Fixed Income'},
            20: {'asset_code': 'FIXEDINCOMEHEADER', 'context': 'Fixed Income (Header Row - produces blank column)', 'description': 'OTPP Investments, Fair Value, Fixed Income Header', 'cost_description': 'OTPP Investments, Cost, Fixed Income Header'},
            21: {'asset_code': 'BONDS', 'context': 'Bonds', 'description': 'OTPP Investments, Fair Value, Bonds', 'cost_description': 'OTPP Investments, Cost, Bonds'},
            22: {'asset_code': 'SHORTTERMINVESTMENTS', 'context': 'Short-term investments', 'description': 'OTPP Investments, Fair Value, Short-term investments', 'cost_description': 'OTPP Investments, Cost, Short-term investments'},
            23: {'asset_code': 'DOMESTICREALRATEPRODUCTS', 'context': 'Canadian real-rate products', 'description': 'OTPP Investments, Fair Value, Canada Real-rate products', 'cost_description': 'OTPP Investments, Cost, Canada Real-rate products'},
            24: {'asset_code': 'FOREIGNREALRATEPRODUCTS', 'context': 'Non-Canadian real-rate products', 'description': 'OTPP Investments, Fair Value, Foreign Real-rate products', 'cost_description': 'OTPP Investments, Cost, Foreign Real-rate products'},
            
            # === OTHER CATEGORIES (Corrected with +10 Row Offset) ===
            27: {'asset_code': 'ALTERNATIVES', 'context': 'Alternative investments', 'description': 'OTPP Investments, Fair Value, Alternative Investments', 'cost_description': 'OTPP Investments, Cost, Alternative Investments'},
            28: {'asset_code': 'INFLATIONSENSITIVEHEADER', 'context': 'Inflation sensitive (Header Row - produces blank column)', 'description': 'OTPP Investments, Fair Value, Inflation Sensitive Header', 'cost_description': 'OTPP Investments, Cost, Inflation Sensitive Header'},
            29: {'asset_code': 'COMMODITIES', 'context': 'Commodities', 'description': 'OTPP Investments, Fair Value, Commodities', 'cost_description': 'OTPP Investments, Cost, Commodities'},
            30: {'asset_code': 'TIMBERLAND', 'context': 'Timberland', 'description': 'OTPP Investments, Fair Value, Timberland', 'cost_description': 'OTPP Investments, Cost, Timberland'},
            31: {'asset_code': 'NATURALRESOURCES', 'context': 'Natural resources', 'description': 'OTPP Investments, Fair Value, Natural Resources', 'cost_description': 'OTPP Investments, Cost, Natural Resources'},
            32: {'asset_code': 'INFLATIONSENSITIVE', 'context': 'Inflation Sensitive Total (after Natural Resources)', 'description': 'OTPP Investments, Fair Value, Inflation Sensitive', 'cost_description': 'OTPP Investments, Cost, Inflation Sensitive'},
            36: {'asset_code': 'REALASSETS', 'context': 'Real Assets Total (after Infrastructure)', 'description': 'OTPP Investments, Fair Value, Real Assets', 'cost_description': 'OTPP Investments, Cost, Real Assets'},
            33: {'asset_code': 'REALASSETSHEADER', 'context': 'Real assets (Header Row - produces blank column)', 'description': 'OTPP Investments, Fair Value, Real Assets Header', 'cost_description': 'OTPP Investments, Cost, Real Assets Header'},
            34: {'asset_code': 'REALESTATE', 'context': 'Real estate', 'description': 'OTPP Investments, Fair Value, Real Estate', 'cost_description': 'OTPP Investments, Cost, Real Estate'},
            35: {'asset_code': 'INFRASTRUCTURE', 'context': 'Infrastructure', 'description': 'OTPP Investments, Fair Value, Infrastructure', 'cost_description': 'OTPP Investments, Cost, Infrastructure'},
            37: {'asset_code': 'TOTAL', 'context': 'Total Investments', 'description': 'OTPP Investments, Fair Value, Total Investments', 'cost_description': 'OTPP Investments, Cost, Total Investments'},

            # === INVESTMENT-RELATED RECEIVABLES SECTION ===
            42: {'asset_code': 'INVESTMENTRELATEDRECEIVABLES', 'context': 'Investment-related receivables (Header)', 'description': 'OTPP Investments, Fair Value, Investment-related receivables', 'cost_description': 'OTPP Investments, Cost, Investment-related receivables'},
            43: {'asset_code': 'SECURITIESREPURCHASED', 'context': 'Securities purchased under agreements to resell', 'description': 'OTPP Investments, Fair Value, Securities purchased under agreements to resell', 'cost_description': 'OTPP Investments, Cost, Securities purchased under agreements to resell'},
            44: {'asset_code': 'CASHCOLLATERALDEPOSITEDUNDERSECURITIES', 'context': 'Cash collateral deposited under securities borrowing', 'description': 'OTPP Investments, Fair Value, Cash collateral deposited under securities', 'cost_description': 'OTPP Investments, Cost, Cash collateral deposited under securities'},
            45: {'asset_code': 'CASHCOLLATERALPAIDUNDERCREDIT', 'context': 'Cash collateral paid under credit support annexes', 'description': 'OTPP Investments, Fair Value, Cash collateral paid under credit', 'cost_description': 'OTPP Investments, Cost, Cash collateral paid under credit'},
            46: {'asset_code': 'DERIVATIVES', 'context': 'Derivative-related, net (receivables)', 'description': 'OTPP Investments, Fair Value, Derivatives', 'cost_description': 'OTPP Investments, Cost, Derivatives'},
            47: {'asset_code': 'INVESTMENTRELATEDSECURITIES', 'context': 'Investment-related receivables total', 'description': 'OTPP Investments, Fair Value, Investment-related receivables total', 'cost_description': 'OTPP Investments, Cost, Investment-related receivables total'},
            
            # === TOTAL INVESTMENTS ===
            48: {'asset_code': 'TOTALINVESTMENTS', 'context': 'Total investments', 'description': 'OTPP Investments, Fair Value, Total investments', 'cost_description': 'OTPP Investments, Cost, Total investments'},
            
            # === INVESTMENT-RELATED LIABILITIES SECTION ===
            49: {'asset_code': 'INVESTMENTRELATEDLIABILITIES', 'context': 'Investment-related liabilities (Header)', 'description': 'OTPP Investments, Fair Value, Investment-related liabilities', 'cost_description': 'OTPP Investments, Cost, Investment-related liabilities'},
            50: {'asset_code': 'SECURITIESSOLD', 'context': 'Securities sold under agreements to repurchase', 'description': 'OTPP Investments, Fair Value, Securities sold under agreements to repurchase', 'cost_description': 'OTPP Investments, Cost, Securities sold under agreements to repurchase'},
            51: {'asset_code': 'SECURITIESSOLDNOTREPURCHASEDLIABILITIES', 'context': 'Securities sold but not yet purchased (Header)', 'description': 'OTPP Investments, Fair Value, Securities sold but not yet purchased', 'cost_description': 'OTPP Investments, Cost, Securities sold but not yet purchased'},
            52: {'asset_code': 'EQUITIESLIABILITIES', 'context': 'Equities under Securities sold but not yet purchased', 'description': 'OTPP Investments, Fair Value, Equities liabilities', 'cost_description': 'OTPP Investments, Cost, Equities liabilities'},
            53: {'asset_code': 'FIXEDINCOMELIABILITIES', 'context': 'Fixed income under Securities sold but not yet purchased', 'description': 'OTPP Investments, Fair Value, Fixed income liabilities', 'cost_description': 'OTPP Investments, Cost, Fixed income liabilities'},
            54: {'asset_code': 'COMMERCIALPAPER1', 'context': 'Commercial paper (first occurrence)', 'description': 'OTPP Investments, Fair Value, Commercial paper 1', 'cost_description': 'OTPP Investments, Cost, Commercial paper 1'},
            55: {'asset_code': 'COMMERCIALPAPERLIABILITIES', 'context': 'Commercial paper (second occurrence)', 'description': 'OTPP Investments, Fair Value, Commercial paper liabilities', 'cost_description': 'OTPP Investments, Cost, Commercial paper liabilities'},
            56: {'asset_code': 'TERMDEBTLIABILITIES', 'context': 'Term debt', 'description': 'OTPP Investments, Fair Value, Term debt liabilities', 'cost_description': 'OTPP Investments, Cost, Term debt liabilities'},
            57: {'asset_code': 'CASHCOLLATERALUNDERSUPPORTLIABILITIES', 'context': 'Cash collateral received under credit support annexes', 'description': 'OTPP Investments, Fair Value, Cash collateral under support liabilities', 'cost_description': 'OTPP Investments, Cost, Cash collateral under support liabilities'},
            58: {'asset_code': 'DERIVATIVESLIABILITIES', 'context': 'Derivative-related, net (liabilities)', 'description': 'OTPP Investments, Fair Value, Derivative-related net liabilities', 'cost_description': 'OTPP Investments, Cost, Derivative-related net liabilities'},
            59: {'asset_code': 'INVESTMENTRELATEDLIABILITIESTOTAL', 'context': 'Investment-related liabilities total', 'description': 'OTPP Investments, Fair Value, Investment-related liabilities total', 'cost_description': 'OTPP Investments, Cost, Investment-related liabilities total'},
            
            # === NET INVESTMENTS ===
            60: {'asset_code': 'NETINVESTMENTS', 'context': 'Net investments (final total)', 'description': 'OTPP Investments, Fair Value, Net investments', 'cost_description': 'OTPP Investments, Cost, Net investments'},
            
            # === MISSING ASSET CODES FOR LIABILITY SECTIONS ===
            61: {'asset_code': 'NETINVESTMENTSLIABILITIES', 'context': 'Net investments liabilities', 'description': 'OTPP Investments, Fair Value, Net investments liabilities', 'cost_description': 'OTPP Investments, Cost, Net investments liabilities'},
            62: {'asset_code': 'REALESTATELIABILITIES', 'context': 'Real estate liabilities', 'description': 'OTPP Investments, Fair Value, Real estate liabilities', 'cost_description': 'OTPP Investments, Cost, Real estate liabilities'},
            63: {'asset_code': 'SECURITIESREPURCHASEDLIABILITIES', 'context': 'Securities repurchased liabilities', 'description': 'OTPP Investments, Fair Value, Securities repurchased liabilities', 'cost_description': 'OTPP Investments, Cost, Securities repurchased liabilities'}
        }
        
        # Column structure based on your source file
        self.column_structure = {
            'B': {'period': '2025-Q2', 'type': 'Fair_Value'},
            'C': {'period': '2025-Q2', 'type': 'Cost'},
            'D': {'period': '2024-Q4', 'type': 'Fair_Value'},
            'E': {'period': '2024-Q4', 'type': 'Cost'}
        }
    
    def scan_and_select_files(self):
        print("=== SCANNING DIRECTORY FOR EXCEL FILES ===")
        excel_files = [f for f in os.listdir('.') if f.lower().endswith(('.xlsx', '.xls'))]
        if not excel_files:
            print("[ERROR] No Excel files found in current directory!")
            return None
        print(f"Found {len(excel_files)} Excel file(s):")
        for i, file in enumerate(excel_files, 1):
            size_kb = os.path.getsize(file) / 1024
            print(f"  {i}. {file} ({size_kb:.1f} KB)")
        if len(excel_files) == 1:
            selected_file = excel_files[0]
            print(f"\n[AUTO] Auto-selected: {selected_file}")
            return selected_file
        while True:
            try:
                choice = int(input(f"\nSelect file (1-{len(excel_files)}): "))
                if 1 <= choice <= len(excel_files):
                    return excel_files[choice - 1]
                else:
                    print(f"Please enter a number between 1 and {len(excel_files)}")
            except ValueError:
                print("Please enter a valid number")
    
    def validate_source_file(self, file_path):
        try:
            print(f"\n=== VALIDATING SOURCE FILE: {file_path} ===")
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            if 'Financial_Schedules' not in workbook.sheetnames:
                print("[ERROR] Missing 'Financial_Schedules' sheet")
                return False
            worksheet = workbook['Financial_Schedules']
            # Updated validation points with the +10 row offset
            validation_points = [
                {'cell': 'A12', 'expected_content': 'Equity', 'description': 'Equity section header'},
                {'cell': 'A14', 'expected_content': 'Canadian', 'description': 'Canadian under Publicly traded'},
                {'cell': 'B14', 'expected_type': 'number', 'description': 'Fair Value data'},
                {'cell': 'A21', 'expected_content': 'Bonds', 'description': 'Bonds under Fixed income'},
            ]
            print("Structure validation:")
            all_valid = True
            for point in validation_points:
                cell_value = worksheet[point['cell']].value
                if 'expected_content' in point:
                    # Loosened check to just see if it contains the word, ignores extra spaces
                    if cell_value and point['expected_content'].lower() in str(cell_value).lower():
                        print(f"  [OK] {point['cell']}: Found '{cell_value}' ({point['description']})")
                    else:
                        print(f"  [ERROR] {point['cell']}: Expected '{point['expected_content']}', found '{cell_value}'")
                        all_valid = False
                elif point['expected_type'] == 'number':
                    is_valid_number = False
                    if isinstance(cell_value, (int, float)):
                        is_valid_number = True
                    elif isinstance(cell_value, str):
                        try:
                            float(cell_value.replace(',', ''))
                            is_valid_number = True
                        except (ValueError, TypeError):
                            is_valid_number = False
                    
                    if is_valid_number:
                        print(f"  [OK] {point['cell']}: Found number-like value '{cell_value}' ({point['description']})")
                    else:
                        print(f"  [ERROR] {point['cell']}: Expected number, found '{cell_value}'")
                        all_valid = False
            return all_valid
        except Exception as e:
            print(f"[ERROR] Validation error: {e}")
            return False
    
    def extract_data(self, file_path):
        try:
            print(f"\n=== EXTRACTING DATA USING CORRECTED MAPPING ===")
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            worksheet = workbook['Financial_Schedules']
            extracted_data = []
            for row_num, mapping in self.row_mappings.items():
                asset_code = mapping['asset_code']
                context = mapping['context']
                print(f"\nRow {row_num:3d}: {asset_code} ({context})")
                # Enhanced extraction with Fair Value = Cost fallback logic
                for col_letter, col_info in self.column_structure.items():
                    period = col_info['period']
                    value_type = col_info['type']
                    cell_value = worksheet[f'{col_letter}{row_num}'].value
                    cleaned_value = self.clean_value(cell_value)
                    
                    # Fair Value = Cost fallback logic for liability items
                    if cleaned_value is None and value_type == 'Fair_Value':
                        # If Fair Value is missing, try to use Cost value (especially for liabilities)
                        cost_col = 'C' if col_letter == 'B' else 'E'  # B->C, D->E
                        cost_value = worksheet[f'{cost_col}{row_num}'].value
                        cost_cleaned = self.clean_value(cost_value)
                        if cost_cleaned is not None:
                            cleaned_value = cost_cleaned
                            print(f"  {col_letter} ({period} {value_type:10s}): {cleaned_value} (using Cost as Fair Value)")
                    
                    # Cost = Fair Value fallback logic for some asset items
                    elif cleaned_value is None and value_type == 'Cost':
                        # If Cost is missing, try to use Fair Value (for certain asset types)
                        fv_col = 'B' if col_letter == 'C' else 'D'  # C->B, E->D
                        fv_value = worksheet[f'{fv_col}{row_num}'].value
                        fv_cleaned = self.clean_value(fv_value)
                        if fv_cleaned is not None:
                            cleaned_value = fv_cleaned
                            print(f"  {col_letter} ({period} {value_type:10s}): {cleaned_value} (using Fair Value as Cost)")
                    
                    if cleaned_value is not None:
                        suffix = '.1' if value_type == 'Fair_Value' else '.2'
                        level = 'REPORTED' if asset_code in ['BONDS', 'COMMODITIES'] else 'NONE'
                        time_series_code = f"ONTARIOTEACHERS.{asset_code}.LEVEL.{level}.H{suffix}@ONTARIOTEACHERS"
                        extracted_data.append({
                            'Time_Series_Code': time_series_code,
                            'Period': period,
                            'Value': cleaned_value
                        })
                        if 'using Cost as Fair Value' not in locals():
                            print(f"  {col_letter} ({period} {value_type:10s}): {cleaned_value}")
                    else:
                        print(f"  {col_letter} ({period} {value_type:10s}): No data (Correctly creates blank cell)")
            print(f"\n[OK] Extraction completed: {len(extracted_data)} data points extracted")
            return extracted_data
        except Exception as e:
            print(f"[ERROR] Extraction failed: {e}")
            return None
    
    def clean_value(self, value):
        if value is None or value == '': return None
        if isinstance(value, str):
            value = value.strip().replace(',', '')
            if value.startswith('(') and value.endswith(')'):
                value = '-' + value[1:-1]
            try: return float(value)
            except ValueError: return None
        return float(value)
    
    def create_output_file(self, extracted_data, output_file):
        # Generate column order to match manual extraction (skip header rows)
        def generate_column_order():
            # Use the exact same order as manual extraction - only data rows, no headers
            data_asset_codes = [
                'EQUITY', 'EQUITIES', 'DOMESTICEQUITIES', 'FOREIGNEQUITIES', 'PRIVATEEQUITY', 
                'DOMESTICPRIVATEEQUITY', 'FOREIGNPRIVATEEQUITY', 'FIXEDINCOME', 'BONDS', 
                'SHORTTERMINVESTMENTS', 'DOMESTICREALRATEPRODUCTS', 'FOREIGNREALRATEPRODUCTS', 
                'ALTERNATIVES', 'INFLATIONSENSITIVE', 'COMMODITIES', 'TIMBERLAND', 
                'NATURALRESOURCES', 'REALASSETS', 'REALESTATE', 'INFRASTRUCTURE',
                'INVESTMENTRELATEDSECURITIES', 'SECURITIESREPURCHASED', 'CASHCOLLATERALPAIDUNDERCREDIT',
                'CASHCOLLATERALDEPOSITEDUNDERSECURITIES', 'DERIVATIVES', 'TOTAL',
                'INVESTMENTRELATEDLIABILITIES', 'SECURITIESSOLD', 'SECURITIESSOLDNOTREPURCHASEDLIABILITIES',
                'EQUITIESLIABILITIES', 'FIXEDINCOMELIABILITIES', 'COMMERCIALPAPER1',
                'COMMERCIALPAPERLIABILITIES', 'TERMDEBTLIABILITIES', 'CASHCOLLATERALUNDERSUPPORTLIABILITIES', 
                'DERIVATIVESLIABILITIES', 'NETINVESTMENTS'
            ]
            
            columns = []
            # Fair Value columns (.1)
            for asset_code in data_asset_codes:
                level = 'REPORTED' if asset_code in ['BONDS', 'COMMODITIES'] else 'NONE'
                columns.append(f"ONTARIOTEACHERS.{asset_code}.LEVEL.{level}.H.1@ONTARIOTEACHERS")
            # Cost columns (.2)  
            for asset_code in data_asset_codes:
                level = 'REPORTED' if asset_code in ['BONDS', 'COMMODITIES'] else 'NONE'
                columns.append(f"ONTARIOTEACHERS.{asset_code}.LEVEL.{level}.H.2@ONTARIOTEACHERS")
            return columns
        
        def generate_descriptions():
            data_asset_codes = [
                'EQUITY', 'EQUITIES', 'DOMESTICEQUITIES', 'FOREIGNEQUITIES', 'PRIVATEEQUITY', 
                'DOMESTICPRIVATEEQUITY', 'FOREIGNPRIVATEEQUITY', 'FIXEDINCOME', 'BONDS', 
                'SHORTTERMINVESTMENTS', 'DOMESTICREALRATEPRODUCTS', 'FOREIGNREALRATEPRODUCTS', 
                'ALTERNATIVES', 'INFLATIONSENSITIVE', 'COMMODITIES', 'TIMBERLAND', 
                'NATURALRESOURCES', 'REALASSETS', 'REALESTATE', 'INFRASTRUCTURE',
                'INVESTMENTRELATEDSECURITIES', 'SECURITIESREPURCHASED', 'CASHCOLLATERALPAIDUNDERCREDIT',
                'CASHCOLLATERALDEPOSITEDUNDERSECURITIES', 'DERIVATIVES', 'TOTAL',
                'INVESTMENTRELATEDLIABILITIES', 'SECURITIESSOLD', 'SECURITIESSOLDNOTREPURCHASEDLIABILITIES',
                'EQUITIESLIABILITIES', 'FIXEDINCOMELIABILITIES', 'COMMERCIALPAPER1',
                'COMMERCIALPAPERLIABILITIES', 'TERMDEBTLIABILITIES', 'CASHCOLLATERALUNDERSUPPORTLIABILITIES', 
                'DERIVATIVESLIABILITIES', 'NETINVESTMENTS'
            ]
            
            # Find description by asset code
            def find_description(asset_code, is_cost=False):
                for mapping in self.row_mappings.values():
                    if mapping['asset_code'] == asset_code:
                        return mapping['cost_description'] if is_cost else mapping['description']
                return f'OTPP Investments, {"Cost" if is_cost else "Fair Value"}, {asset_code}'
            
            descriptions = []
            # Fair Value descriptions
            for asset_code in data_asset_codes:
                descriptions.append(find_description(asset_code, False))
            # Cost descriptions
            for asset_code in data_asset_codes:
                descriptions.append(find_description(asset_code, True))
            return descriptions
        
        fixed_column_order = generate_column_order()
        fixed_header_descriptions = generate_descriptions()
        
        
        try:
            print(f"\n=== CREATING OUTPUT FILE WITH HARDCODED HEADERS: {output_file} ===")
            if not extracted_data:
                print("[ERROR] No data to export")
                return False
            
            df = pd.DataFrame(extracted_data)
            pivot_df = df.pivot(index='Period', columns='Time_Series_Code', values='Value').reindex(columns=fixed_column_order)
            pivot_df = pivot_df.sort_index(ascending=False)
            
            final_data = [[''] + fixed_column_order, [''] + fixed_header_descriptions]
            for period, row in pivot_df.iterrows():
                final_data.append([period] + list(row.values))

            final_df = pd.DataFrame(final_data)
            
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                final_df.to_excel(writer, sheet_name='DATA', index=False, header=False)
                worksheet = writer.sheets['DATA']
                from openpyxl.styles import Font
                bold_font = Font(bold=True)
                for col_idx, col_cells in enumerate(worksheet.columns):
                    max_len = max(len(str(cell.value or '')) for cell in col_cells)
                    width = 15 if col_idx == 0 else min(max_len + 2, 50)
                    worksheet.column_dimensions[col_cells[0].column_letter].width = width
                for col in range(1, len(fixed_column_order) + 2):
                    worksheet.cell(row=1, column=col).font = bold_font
                    worksheet.cell(row=2, column=col).font = bold_font
            
            print(f"[OK] Output file created successfully.")
            print(f"[OK] Format: Two hardcoded header rows with periods as data rows.")
            print(f"[OK] Time series: {len(fixed_column_order)}")
            print(f"[OK] Periods: {len(pivot_df.index)} ({', '.join(pivot_df.index)})")
            return True
        except Exception as e:
            print(f"[ERROR] Error creating output: {e}")
            import traceback
            traceback.print_exc()
            return False

def main():
    print("=" * 80)
    print("   OTPP HIERARCHICAL DATA EXTRACTOR (V10 - Row Offset Fix)")
    print("=" * 80)
    
    extractor = OTTPHierarchicalExtractor()
    print(f"[OK] Loaded {len(extractor.row_mappings)} row mappings for 74-column output.")
    
    source_file = extractor.scan_and_select_files()
    if not source_file: return
    
    if not extractor.validate_source_file(source_file):
        print("\n[ERROR] Source file validation failed! Please check the file and try again.")
        return
    
    print(f"\n[OK] Source file validated: {source_file}")
    
    extracted_data = extractor.extract_data(source_file)
    
    if extracted_data:
        output_file = f"CANF_OTPP_DATA_{datetime.now().strftime('%Y%m%d')}.xlsx"
        if extractor.create_output_file(extracted_data, output_file):
            print(f"\n[SUCCESS] SUCCESS! Created: {output_file}")
        else:
            print(f"\n[ERROR] Failed to create the output file.")
    else:
        print(f"\n[ERROR] No data was extracted from the source file.")

if __name__ == "__main__":
    main()