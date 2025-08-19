import os
import glob
from pathlib import Path
from typing import List, Optional
import fitz  # PyMuPDF
from pdf2docx import Converter
from docx import Document
from docx.document import Document as DocumentClass
from docx.oxml.table import CT_Tbl
from docx.oxml.text.paragraph import CT_P
from docx.table import Table
from docx.text.paragraph import Paragraph
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# --- Core Logic ---

def extract_intro_text(pdf_path: str, page_num: int, start_marker: str, end_marker: str) -> Optional[str]:
    """Extracts the introductory text from a specific page of the PDF."""
    try:
        doc = fitz.open(pdf_path)
        if page_num >= len(doc): return None
        page = doc[page_num]
        text = page.get_text("text")
        doc.close()
        start_index = text.find(start_marker)
        if start_index == -1: return None
        end_index = text.find(end_marker, start_index)
        if end_index == -1: return None
        intro = text[start_index:end_index].strip().replace('NOTE 2.\n', 'NOTE 2. ').replace('\n', ' ')
        return intro
    except Exception as e:
        print(f"      - WARN: Could not extract intro text: {e}")
        return None

def stitch_and_clean_content_from_docx(doc: DocumentClass) -> List[List[str]]:
    """
    Iterates through all elements, stitches them, and performs intelligent cleaning
    to remove conversion artifacts like duplicated data rows.
    """
    stitched_rows = []
    
    # First Pass: Stitch all paragraphs and table rows together
    for block in doc.element.body:
        if isinstance(block, CT_P):
            para = Paragraph(block, doc)
            text = para.text
            if text.strip():
                stitched_rows.append([text]) # Keep paragraph text
        elif isinstance(block, CT_Tbl):
            table = Table(block, doc)
            for row in table.rows:
                # Preserve indentation in the first column, strip others
                new_row = [cell.text if i == 0 else cell.text.strip() for i, cell in enumerate(row.cells)]
                cleaned_row = [cell for cell in new_row if str(cell).strip() != '$']
                stitched_rows.append(cleaned_row)
            
    # Second Pass: Intelligent Duplicate Removal
    cleaned_final_rows = []
    previous_numeric_data = None
    for row in stitched_rows:
        is_data_row = len(row) > 1 and row[0] and str(row[0]).strip() and any(c.replace(',','').replace('.','').isdigit() for c in row[1:])

        if is_data_row:
            current_numeric_data = tuple(str(cell).strip() for cell in row[1:])
            if current_numeric_data == previous_numeric_data:
                print(f"    - Found and removed a duplicate artifact row for label: '{row[0]}'")
                continue
            previous_numeric_data = current_numeric_data
        else:
            previous_numeric_data = None
            
        cleaned_final_rows.append(row)
        
    return cleaned_final_rows

def write_data_to_excel(intro_text: Optional[str], all_rows: List[List[str]], output_path: Path):
    """Writes the intro text and all stitched rows to an Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Financial_Schedules"

    current_row = 1
    if intro_text:
        max_cols = max(len(row) for row in all_rows if row) if all_rows else 5
        intro_cell = ws.cell(row=current_row, column=1, value=intro_text)
        intro_cell.alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + 2, end_column=max_cols)
        current_row += 4

    for r_idx_offset, row_data in enumerate(all_rows):
        row_idx = current_row + r_idx_offset
        for c_idx, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=c_idx, value=cell_value)
            
            row_text_lower = " ".join(str(v).strip().lower() for v in row_data if v)
            is_header = 'fair value' in row_text_lower or 'as at' in row_text_lower or 'level 1' in row_text_lower or 'total' in row_text_lower
            is_category_header = (len([v for v in row_data if v and str(v).strip()]) == 1 and c_idx == 1)
            is_total_row = (not row_data[0] or not str(row_data[0]).strip()) and (len(row_data) > 1 and row_data[1] and str(row_data[1]).strip())

            if is_header or is_category_header or is_total_row:
                cell.font = Font(bold=True)

    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 25

    wb.save(output_path)

def main():
    """Main function to orchestrate the full PDF -> DOCX -> Multi-Sheet Excel workflow."""
    target_directory = '.'
    
    print(f"[*] Scanning for PDF files in '{target_directory}'...")
    pdf_files = glob.glob(os.path.join(target_directory, '**', '*.pdf'), recursive=True)
    
    if not pdf_files:
        print("\nNo PDF files found.")
        return

    print(f"[+] Found {len(pdf_files)} PDF file(s).")
    
    for pdf_path_str in pdf_files:
        pdf_path = Path(pdf_path_str)
        print(f"\n--- Processing: {pdf_path} ---")
        temp_docx_path = Path(f"{pdf_path.stem}_temp.docx")
        output_excel_path = Path(f"{pdf_path.stem}_extracted.xlsx")
        
        try:
            print("  - Step 1: Extracting introductory text...")
            intro_text = extract_intro_text(pdf_path_str, 5, "NOTE 2.", "(Canadian $ millions)")
            print(f"  - Step 1: Text extraction {'successful' if intro_text else 'failed'}.")

            print("  - Step 2: Converting relevant PDF pages (6-8) to a temporary DOCX...")
            cv = Converter(pdf_path_str)
            cv.convert(str(temp_docx_path), start=5, end=8)
            cv.close()
            print("  - Step 2: Conversion successful.")
            
            print("  - Step 3: Stitching and cleaning all content from the DOCX...")
            doc = Document(temp_docx_path)
            final_rows = stitch_and_clean_content_from_docx(doc)
            print(f"    - Stitched and cleaned {len(final_rows)} total rows of content.")
            
            if final_rows:
                print(f"  - Step 4: Writing final data to '{output_excel_path}'...")
                write_data_to_excel(intro_text, final_rows, output_excel_path)
                print(f"  - Step 4: SUCCESS! Excel file created.")
            else:
                print("    - FAILED: No content remained after cleaning.")

        except Exception as e:
            print(f"  - ERROR: A critical error occurred: {e}")
        finally:
            if os.path.exists(temp_docx_path):
                os.remove(temp_docx_path)
                print("  - Step 5: Cleaned up temporary DOCX file.")

    print("\n[+] Workflow finished.")

if __name__ == '__main__':
    main()